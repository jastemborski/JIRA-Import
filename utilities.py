import openpyxl
import requests
import constants
import json
import getpass
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import PatternFill
from ADT import ProjectMeta, IssueTypeMeta, Story, Issue
from JiraApi import *

info_row = 2
PROCESS_DICT = {'Complaint': 'B', 'Inquiry': 'C', 'CAPA': 'D',
                'Quality Event': 'E', 'Quality Event Investigation': 'E',
                'Change Control': 'F', 'Change Control Plan': 'F',
                'Assessment': 'G', 'Investigation/Evaluation': 'H',
                'Audit': 'I', 'Audit Findings': 'I', 'Audit Plan': 'I',
                'Products': 'J', 'Contacts': 'K',
                'Notes': 'L', 'Tasks': 'M', 'Global': 'N',
                }

COL_DICT = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H',
            9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O',
            16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V',
            23: 'W', 24: 'X', 25: 'Y', 26: 'Z'}


""" =========================================================================
|                           Accessor Methods                                |
=========================================================================="""


def get_change_type(req_sheet, row_num):
    """ Accessor for Change Type
    Args:
        req_sheet: A variable holding an Excel Workbook sheet in memory.
        row_num: A variable holding the row # of the data being accessed.
    Returns:
            A string value of the Change Type
"""
    return (req_sheet['A' + str(row_num)].value)


def get_change_description(req_sheet, row_num):
    """ Accessor for Change Description
    Args:
        req_sheet: A variable holding an Excel Workbook sheet in memory.
        row_num: A variable holding the row # of the data being accessed.
    Returns:
            A string value of the Change Description
    """
    return (req_sheet['B' + str(row_num)].value)


def get_platform(req_sheet, row_num):
    """ Accessor for Platform
    Args:
        req_sheet: A variable holding an Excel Workbook sheet in memory.
        row_num: A variable holding the row # of the data being accessed.
    Returns:
            A string value of the Platform
"""
    return (req_sheet['C' + str(row_num)].value)


def get_process(req_sheet, row_num):
    """ Accessor for Process
    Args:
        req_sheet: A variable holding an Excel Workbook sheet in memory.
        row_num: A variable holding the row # of the data being accessed.
    Returns:
            A string value of the Process
    """
    return (req_sheet['D' + str(row_num)].value)


def get_notes(req_sheet, row_num):
    """ Accessor for Notes
    Args:
        req_sheet: A variable holding an Excel Workbook sheet in memory.
        row_num: A variable holding the row # of the data being accessed.
    Returns:
            A string value of the Notes
    """
    return (req_sheet['E' + str(row_num)].value)


def get_jira_task(req_sheet, row_num):
    """ Accessor for JIRA Key
    Args:
        req_sheet: A variable holding an Excel Workbook sheet in memory.
        row_num: A variable holding the row # of the data being accessed.
    Returns:
            A string value of the Notes
    """
    return (req_sheet['F' + str(row_num)].value)


def get_status(req_sheet, row_num):
    """ Accessor for JIRA Key
    Args:
        req_sheet: A variable holding an Excel Workbook sheet in memory.
        row_num: A variable holding the row # of the data being accessed.
    Returns:
            A string value of the Notes
    """
    return (req_sheet['G' + str(row_num)].value)


def get_subtask_assignee(req_sheet, row_num):
    """ Accessor for the Assignee field
    Args:
        req_sheet: A variable holding an Excel Workbook sheet in memory.
        row_num: A variable holding the row # of the data being accessed.
    Returns:
            A string value of the Assignee
    """
    return (req_sheet['H' + str(row_num)].value)


def get_sprint(req_sheet, row_num):
    """ Accessor for the Sprint field
    Args:
        req_sheet: A variable holding an Excel Workbook sheet in memory.
        row_num: A variable holding the row # of the data being accessed.
    Returns:
            A string value of the Sprint
    """
    return (req_sheet['I' + str(row_num)].value)


def get_comments(req_sheet, row_num):
    """ Accessor for Comments field
    Args:
        req_sheet: A variable holding an Excel Workbook sheet in memory.
        row_num: A variable holding the row # of the data being accessed.
    Returns:
            A string value of the Comments
    """
    return (req_sheet['J' + str(row_num)].value)


def get_customer(customer_sheet):
    """ Accessor for Customer

    Args:
        customer_sheet: A variable holding an Excel Workbook sheet in memory.
    Returns:
            A string value of the Customer
    """
    return (customer_sheet['A' + "2"].value)


def get_parent(jira_sheet, process):
    """ Accessor for Parent

    Accessor method for retrieving the value for Parent (JIRA Key) on the
    JIRA Stories Sheet.

    There is a check to make certain the process in question is amongst those
    qualified to exist.

    Args:
        jira_sheet: A variable holding an Excel Workbook sheet in memory.
        process: A variable holding the process of an Issue.
    Returns:
            A string value of the Parent
"""
    if process in PROCESS_DICT:
        return (jira_sheet[PROCESS_DICT.get(process) + "2"].value)
    else:
        print("""Error: " + process + " is an invalid process.
                The following QE processes are acceptable: Complaints, Inquiry,
                CAPA, Quality Event, Change Control.\n""")


def get_story_assignee(jira_sheet, process):
    """ Accessor for Story Assignee

    Accessor method for retrieving the value for Story Assignee on the
    JIRA Stories Sheet.

    There is a check to make certain the process in question is amongst those
    qualified to exist.

    Args:
        jira_sheet: A variable holding an Excel Workbook sheet in memory.
        process: A variable holding the process of an Issue.
    Returns:
            A string value of the Parent
"""
    if process in PROCESS_DICT:
        return (jira_sheet[PROCESS_DICT.get(process) + "6"].value)
    else:
        print("""Error: " + process + " is an invalid process.
                The following QE processes are acceptable: Complaints, Inquiry,
                CAPA, Quality Event, Change Control.\n""")


def get_stories(jira_sheet, customer_sheet):
    """ Captures information residing in the JIRA Stories sheet of the Workbook

    Populates a list of Story objects corresponding to the data captured on the
    JIRA Stories sheet.

    Args:
        jira_sheet: A variable holding an Excel Workbook sheet in memory.
    Returns:
        story_list: A list of Story objects corresponding to
                  table row data fetched. For example:
                    key: TEST-303
                    summary: CAPA
                    description: CAPA module
                    project: TEST
                    col: D
    """
    story_list = []
    for c in range(2, jira_sheet.max_column):
        key = jira_sheet[COL_DICT[c] + "2"].value
        if key is None:
            key = ""
        # print(key)
        story = Story(key,  # key
                      jira_sheet[COL_DICT[c] + "3"].value,  # summary
                      jira_sheet[COL_DICT[c] + "4"].value,  # description
                      jira_sheet[COL_DICT[c] + "5"].value,  # project
                      customer_sheet['A2'].value,  # customer
                      jira_sheet[COL_DICT[c] + "6"].value,  # assignee
                      jira_sheet[COL_DICT[c] + "7"].value,  # board
                      jira_sheet[COL_DICT[c] + "8"].value,  # sprint
                      COL_DICT[c])  # col
        story_list.append(story)
        # print(story.key)
    return story_list


""" =========================================================================
|                          End Accessor Methods                            |
=========================================================================="""


def create_stories(story_dict, session=None, wb=None, filename=None):
    """ Creates JIRA Stories if not already created.

    Checks to see if the row 'Key', designating the JIRA Story key, has been
    populated - indicating the Story already exists within JIRA.

    If the story doesn't exist and both the 'Description' and 'Title'
    rows have been populated, indicating a Story should be created, - create
    the Stories the stories then update the Excel Workbook.

    Args:
        story_dict: A list of Story objects captured on the JIRA Stories sheet.
        session: A variable capturing the session info previously captured via
                the Login function. ( Optional )
        wb: A variable storing the Excel Workbook in memory. ( Optional )
    """
    try:
        for story in story_dict:
            # print(story.key)
            if not story.key:
                _issue = Issue()
                _issue.customer = story.customer
                _issue.project_key = constants.PROJECT_KEY
                _issue.process = story.summary
                _issue.change_description = story.description
                _issue.board = story.board
                _issue.sprint = story.sprint
                # print(story.assignee)
                _issue.assignee = story.assignee
                if story.summary and story.description:
                    issue = jira_create_issue(_issue, True, session)
                    story_json = json.loads(issue.text)
                    story_key = story_json['key']
                    # print(story_key)
                    # print(story.board)
                    # print(story.sprint)
                    # if story.board is not None:
                    #     print(move_to_sprint(session, story.board,
                    #                          story.sprint, story_key))
                    # print(issue)
                    # print(issue.text)
                    info = json.loads(issue.text)
                    key = info['key']
                    write_story(wb, story.col, key, filename)
    except Exception:
        # raise Exception
        print("Invalid issue\n")


def write_story(wb, col, key, filename):
    """ Writes Stories to Excel Workbook.

    Args:
        wb: A variable storing the Excel Workbook in memory.
        col: A variable containing the column being updated.
        key: A variable containing the JIRA Story Key.
    """
    try:
        jira_sheet = wb.get_sheet_by_name('JIRA Stories')
        jira_sheet[col + "2"] = key
        wb.save(filename)
    except Exception:
        print("""Unable to save workbook. Please close excel spreadsheet then
               try again.""")


def readFile(filename=None):
    """ Reads Excel Workbook into memory

    Args:
        filename: A variable holding the document's filename

    """
    try:
        if filename is None:
            filename = input("Please enter a filename: ")
        wb = openpyxl.load_workbook(filename=filename)
        return wb
    except Exception:
        print("Invalid file name\n")


def parseFile(wb, session=None, filename=None):
    """ Parses Workbook into Issues Objects

        Iterates through each row containing data within the Workbook
        while creating an Issue object capturing all of the information.

        A Story ( Parent ) is required in order to create Sub-tasks
        ( info on Requirements sheet) therefore there is a check prior to
        processing the file in place in order to create a Story ( Parent )
        if they don't already exist.

    Args:
        wb: A variable holding a Excel Workbook in memory.
        session: An optional variable holding the current Session.
        filename: An optional variable holding the filename of the Workbook
    Return: A list of Issue Objects
    """
    try:
        req_sheet = wb.get_sheet_by_name('Requirements')
        jira_sheet = wb.get_sheet_by_name('JIRA Stories')
        customer_sheet = wb.get_sheet_by_name('Customer Info')
        issues = []
        maxRow = req_sheet.max_row
        customer = get_customer(customer_sheet)

        # Determines if Stories have already been created for each Sub-task
        stories = get_stories(jira_sheet, customer_sheet)
        create_stories(stories, session, wb, filename)

        #  Process File
        for row in range(2, maxRow + 1):
            issue = Issue()
            issue.change_type = get_change_type(req_sheet, row)
            issue.change_description = get_change_description(req_sheet, row)
            issue.platform = str(get_platform(req_sheet, row))
            issue.process = str(get_process(req_sheet, row))
            issue.notes = get_notes(req_sheet, row)
            issue.parent = get_parent(jira_sheet, issue.process)
            issue.customer = customer
            issue.issue_type = constants.SUBTASK
            issue.project_key = constants.PROJECT_KEY
            issue.jira_key = get_jira_task(req_sheet, row)
            issue.row = str(row)
            issue.assignee = get_subtask_assignee(req_sheet, row)
            issues.append(issue)
        return issues
    except Exception:
        print("Error processing file.")


def login():
    """ Method for logging into JIRA and capturing Session
        Args: None
        Return: Session
    """
    authenticated = False
    while authenticated is not True:
        username = input("Please enter your JIRA username: ")
        password = getpass.getpass("Please enter your JIRA password: ")
        session = requests.Session()
        cred = json.dumps({"username": username, "password": password})
        s = session.post(constants.URI_LOGIN,
                         headers=constants.APPLICATION_JSON,
                         data=cred)
        authenticated = s.ok
        if (not s.ok):
            print("\nInvalid login credentials. \n \
                    Please enter the correct username/password. \n")
        # print(s.status_code)
    return session


def get_issuetypes(json_createmeta):
    issue_types = json.loads(json_createmeta)
    num_projects = len(issue_types["projects"])
    metadata = []
    for i in range(0, num_projects):
        p_url = issue_types["projects"][i]['self']
        p_id = issue_types["projects"][i]['id']
        p_key = issue_types["projects"][i]['key']
        p_name = issue_types["projects"][i]['name']
        num_issues = len(issue_types["projects"][i]["issuetypes"])
        p_data = ProjectMeta(p_url, p_id, p_key, p_name)
        metadata.insert(i, p_data)
        # print(p_data)
        for x in range(0, num_issues):
            i_url = issue_types["projects"][i]['issuetypes'][x]['self']
            i_id = issue_types["projects"][i]['issuetypes'][x]['id']
            i_desc = issue_types["projects"][i]['issuetypes'][x]['description']
            i_name = issue_types["projects"][i]['issuetypes'][x]['name']
            task_metadata = IssueTypeMeta(i_url, i_id, i_desc, i_name)
            metadata[i].issuetype.append(task_metadata)
            # print(task_metadata)
    return metadata


def write_jira_key(issues, filename):
    try:
        wb = openpyxl.load_workbook(filename)
        req_sheet = wb.get_sheet_by_name('Requirements')
        for issue in issues:
            req_sheet['F' + str(issue.row)] = issue.jira_key
            val = constants.URL_BROWSE + issue.jira_key
            req_sheet['F' + str(issue.row)].hyperlink = Hyperlink(ref="",
                                                                  target=val)
        wb.save(filename)
    except Exception:
        print("Error: Write Jira Key - Please close the file, then try again.")


def update_status(issue_list, session, filename):
    try:
        row_seed = 2
        wb = openpyxl.load_workbook('jira-import-template.xlsx')
        req_sheet = wb.get_sheet_by_name('Requirements')
        for issue in issue_list:
            if issue.jira_key:
                issueStatusJson = get_issue(issue.jira_key, session)
                issueStatus = json.loads(issueStatusJson.text)
                issueStatus = issueStatus['fields']['status']['name']
                req_sheet['G' + str(row_seed)] = issueStatus
                if issueStatus == "To Do":
                    highlight_row(issue.row, "FFC7CE", req_sheet)
                elif issueStatus == "Done":
                    highlight_row(issue.row, "C6EFCE", req_sheet)
                else:
                    highlight_row(issue.row, "FFEB9C", req_sheet)
            row_seed += 1
        wb.save(filename)
    except Exception:
        print("Please close the file, then try again.")


""" Todo: def scrub_doc cleans internal info to send dox to client """
""" Todo: def merge_docs merges clients and local """


def form_query(issue):
    summary = issue.process + ': ' + issue.change_type
    query = "project=" + issue.project_key + " and summary~'" \
            + summary + "' and description~'" + issue.change_description + "'"
    return query


def write_status(issues, filename, session):
    try:
        wb = openpyxl.load_workbook(filename)
        req_sheet = wb.get_sheet_by_name('Requirements')
        for issue in issues:
            req_sheet['G' + issue.row] = issue.status
            if issue.status == "To Do":
                highlight_row(issue.row, "FFC7CE", req_sheet)
            elif issue.status == "Done":
                highlight_row(issue.row, "C6EFCE", req_sheet)
            else:
                highlight_row(issue.row, "FFEB9C", req_sheet)
        wb.save(filename)
    except Exception:
        print("Please close the file, then try again.")


def create_issues(session, issues, filename):
    # non_created_issues = [(issue for issue in Issues
    #                        if not is_issue_created(issue))]
    # for issue in non_created_issues:
    #     print(issue)
    try:
        non_created_issues = []
        for issue in issues:
            duplicate = is_duplicate(issue, session, filename)
            if not issue.jira_key and not duplicate:
                non_created_issues.append(issue)
        if non_created_issues:
            json_issue_response = jira_create_issues(session,
                                                     issues)
            print(json_issue_response)
            i = 0
            for pending_issue in non_created_issues:
                issue_reponse = json.loads(json_issue_response.text)
                pending_issue.jira_key = issue_reponse['issues'][i]['key']
                issueStatusJson = get_issue(pending_issue.jira_key, session)
                issueStatus = json.loads(issueStatusJson.text)
                pending_issue.status = issueStatus['fields']['status']['name']
                i = i + 1
            write_jira_key(non_created_issues, filename)
            write_status(non_created_issues, filename, session)
        else:
            return print("No issues to add!")
    except Exception:
        print("Error: Create_Issues")


def is_issue_created(issue, s):
    return True if issue.jira_key else False


def is_duplicate(issue, session, filename):
    """
        TODO: For some reason JQL doesn't like apostrophes... investigate
        Also need to investigate this method in production
        {'errorMessages': ['Unrecognized field "fieldsByKeys"
        (Class com.atlassian.jira.rest.v2.search.SearchRequestBean),
        not marked as ignorable\n at
        [Source: org.apache.catalina.connector.CoyoteInputStream@713e4dd2;
        line: 1, column: 221] (through reference chain:
        com.atlassian.jira.rest.v2.search.SearchRequestBean["fieldsByKeys"])']}
    """
    try:
        search_query = form_query(issue)
        field_list = []
        field_list.append("summary, status")
        search = json.loads(search_issues(search_query,
                                          field_list=field_list,
                                          session=session).text)
        if search['total'] and not issue.jira_key:
            print(json.dumps(search, indent=3))
            wb = openpyxl.load_workbook(filename)
            req_sheet = wb.get_sheet_by_name('Requirements')
            req_sheet['E' + str(issue.row)] = "Issue: " + \
                search['issues'][0]["key"] + " already exists."
            req_sheet['G' + issue.row] = "Duplicate"
            highlight_row(issue.row, "DBDBDB", req_sheet)
            wb.save(filename)
            return True
        else:
            return False
    except Exception:
        print("Error in issue created")
        print(issue)
        print(search)


def highlight_row(row, hex_color, sheet):
    max_col = sheet.max_column
    color = PatternFill("solid", fgColor=hex_color)
    for col in range(1, max_col + 1):
        sheet[COL_DICT[col] + row].fill = color


def update():
    return False


def retrieve(key, session):
    """
    sprint
    """
    issue_info = json.loads(get_issue(key, session).text)
    # print(json.dumps(issue_info, indent=3))

    # _change_type = extract_change_type(issue_info['fields']['summary'])
    # print(_change_type)
    # _platform = issue_info['fields'][constants.ID_JIRA_PLATFORM]
    # print(_platform)
    # _process = issue_info['fields'][constants.ID_JIRA_PROCESS]
    # print(_process)
    # _change_description = issue_info['fields']['description']
    # print(_change_description)
    # _additional_notes = extract_comments(issue_info['fields']['comment']['comments']
    # print(extract_comments(issue_info['fields']['comment']['comments']))
    # _status = issue_info['fields']['status']['name']
    # print(_status)
    # _assignee = issue_info['fields']['reporter']['displayName']
    # print(_assignee)
    print(issue_info['fields']['customfield_10115'])


def extract_comments(comments):
    """ Utility method for parsing JIRA comments represented as JSON
        Args:
            comments: A variable containing JIRA comments in JSON
                      representation.
        Returns:
            A string containing all of the JIRA comments tied to an issue
    """
    size = len(comments)
    addtional_notes = ""
    for n in range(0, size):
        addtional_notes = addtional_notes + comments[n]['body'] + "\n"
    return addtional_notes


def extract_change_type(summary):
    _summary_list = summary.split(":")
    return (_summary_list[1].strip())


def get_board_id(board_name, session):
    all_boards = json.loads(get_all_boards(session).text)
    num_boards = len(all_boards['values'])
    for i_board in range(0, num_boards):
        if all_boards['values'][i_board]['name'].upper() == board_name.upper():
            return str(all_boards['values'][i_board]['id'])


def get_sprint_id(board_id, sprint_name, session):
    all_sprints = json.loads(get_all_sprints(board_id, session).text)
    num_sprints = len(all_sprints['values'])
    for index in range(0, num_sprints):
        if all_sprints['values'][index]['name'].upper() == sprint_name.upper():
            return str(all_sprints['values'][index]['id'])


def move_to_sprint(session, board_name, sprint_name, issue_key):
    try:
        board_id = get_board_id(board_name, session)
        if board_id is None:
            error_message = "Could not find the board."
            raise TypeError
        sprint_id = get_sprint_id(board_id, sprint_name, session)
        if sprint_id is None:
            error_message = "Could not find the sprint."
            raise TypeError
        return move_issues_to_sprint(sprint_id, issue_key, session)
    except TypeError:
        print(error_message)
